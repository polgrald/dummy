import pandas as pd
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import os
import warnings
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook as _oxl_load_workbook

# Suppress the openpyxl warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


class CustomerEmailProcessor:
    def __init__(self, excel_file_path):
        self.excel_file_path = excel_file_path
        self.customer_data = defaultdict(list)
        # Lightweight persistent preferences
        # Store preferences next to this script so they persist regardless of the current working directory
        try:
            self._prefs_path = (Path(__file__).parent / '.email_prefs.json').resolve()
        except Exception:
            # Fallback to CWD if __file__ is not available for any reason
            self._prefs_path = Path('.email_prefs.json').resolve()
        self._prefs = self._load_prefs()

    # -------------------- Preferences helpers --------------------
    def _load_prefs(self):
        """Load preferences from the script directory, with a fallback migration
        from a .email_prefs.json that may have been created in the current working directory
        by older versions.
        """
        # 1) Primary: script-local prefs file
        try:
            if self._prefs_path.exists():
                import json
                with open(self._prefs_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if isinstance(data, dict):
                        return data
        except Exception:
            pass

        # 2) Fallback/migration: look in the current working directory
        try:
            cwd_prefs_path = (Path.cwd() / '.email_prefs.json').resolve()
            if cwd_prefs_path != self._prefs_path and cwd_prefs_path.exists():
                import json
                with open(cwd_prefs_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if isinstance(data, dict):
                        # Save a copy into the script-local location for future runs
                        try:
                            self._prefs_path.parent.mkdir(parents=True, exist_ok=True)
                            with open(self._prefs_path, 'w', encoding='utf-8') as wf:
                                json.dump(data, wf, indent=2)
                        except Exception:
                            pass
                        return data
        except Exception:
            pass

        return {}

    def _save_pref(self, key, value):
        try:
            import json
            self._prefs[key] = value
            # Ensure parent directory exists
            try:
                self._prefs_path.parent.mkdir(parents=True, exist_ok=True)
            except Exception:
                pass
            with open(self._prefs_path, 'w', encoding='utf-8') as f:
                json.dump(self._prefs, f, indent=2)
        except Exception:
            # Non-fatal if we cannot persist
            pass

    def _get_pref(self, key, default=None):
        return self._prefs.get(key, default)

    def separate_data_per_customer(self):
        """Main function to process customer data and create emails"""
        try:
            # Ensure local copy if on cloud storage / placeholder
            self.excel_file_path = self._ensure_local_copy(self.excel_file_path)
            print("Reading Excel file...")
            print(f"File path: {self.excel_file_path}")

            # First, robustly list sheet names
            try:
                sheet_names = self._list_sheet_names_robust(self.excel_file_path)
                print(f"📊 Available sheets in your file: {sheet_names}")

                if not sheet_names:
                    print(
                        "❌ No sheets detected in the workbook. Please ensure the file is a valid .xlsx/.xlsm and not password-protected.")
                    return

                # Try auto-selecting common sheet names (case-insensitive)
                lowered = {name.lower(): name for name in sheet_names}
                data_sheet_name = None
                emails_sheet_name = None

                for candidate in ("data", "sheet1", "sheet 1", "invoices", "transactions"):
                    if candidate in lowered:
                        data_sheet_name = lowered[candidate]
                        break

                for candidate in ("emails", "email", "contacts", "customers"):
                    if candidate in lowered:
                        emails_sheet_name = lowered[candidate]
                        break

                # If not auto-detected, let the user choose which sheet contains the data
                if not data_sheet_name:
                    data_sheet_name = self._choose_sheet(sheet_names, "data")
                if not data_sheet_name:
                    return

                # Let the user choose which sheet contains emails (optional) if not auto-detected
                if emails_sheet_name is None:
                    emails_sheet_name = self._choose_sheet(sheet_names, "emails", optional=True)

                print(f"✅ Using data sheet: '{data_sheet_name}'")
                if emails_sheet_name:
                    print(f"✅ Using emails sheet: '{emails_sheet_name}'")
                else:
                    print("ℹ️  No emails sheet selected - will prompt for emails later")

            except Exception as e:
                print(f"❌ Error reading Excel file structure: {str(e)}")
                return

            # Read the data sheet (with robust fallbacks)
            try:
                data_sheet = self._read_sheet_df_robust(data_sheet_name)
                print(f"✅ Successfully read data sheet with {len(data_sheet)} rows")
            except Exception as e:
                print(f"❌ Error reading data sheet: {str(e)}")
                return

            # Read the emails sheet if selected
            emails_sheet = pd.DataFrame()
            if emails_sheet_name:
                try:
                    emails_sheet = self._read_sheet_df_robust(emails_sheet_name)
                    print(f"✅ Successfully read emails sheet with {len(emails_sheet)} rows")
                except Exception as e:
                    print(f"⚠️  Could not read emails sheet: {str(e)}")
                    print("Continuing without email addresses...")

            # Show what columns we found
            print(f"📋 Columns in data sheet: {list(data_sheet.columns)}")
            if not emails_sheet.empty:
                print(f"📧 Columns in emails sheet: {list(emails_sheet.columns)}")

            # Validate data exists
            if len(data_sheet) == 0:
                print("❌ No data found in data sheet!")
                return

            # Process customer data
            self._process_customer_data(data_sheet)

            if not self.customer_data:
                print("❌ No customer data found to process.")
                return

            # Create emails for each customer
            self._create_emails(emails_sheet)

            print(f"✅ Process completed! {len(self.customer_data)} customer emails prepared.")

        except Exception as e:
            print(f"❌ Error processing data: {str(e)}")
            import traceback
            traceback.print_exc()

    def _list_sheet_names_robust(self, file_path):
        """Robust method to list sheet names with multiple fallback strategies"""
        file_path = Path(file_path)

        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        print(f"🔍 Attempting to read sheets from: {file_path}")
        print(f"📊 File size: {file_path.stat().st_size} bytes")

        # Strategy 1: Try pandas with multiple engines (without calamine)
        engines_to_try = [
            None,  # Let pandas choose
            'openpyxl',
            'xlrd'
        ]

        # Add pyxlsb only for .xlsb files if available
        if file_path.suffix.lower() == '.xlsb':
            try:
                import pyxlsb
                engines_to_try.append('pyxlsb')
            except ImportError:
                print("ℹ️ pyxlsb not installed for .xlsb files")

        for engine in engines_to_try:
            try:
                print(f"  Trying pandas with engine: {engine}")
                with pd.ExcelFile(file_path, engine=engine) as xl:
                    sheets = xl.sheet_names
                    if sheets:
                        print(f"✅ Success with engine: {engine}")
                        return sheets
            except Exception as e:
                print(f"  ❌ Engine {engine} failed: {str(e)}")
                continue

        # Strategy 2: Try openpyxl directly
        try:
            print("  Trying openpyxl directly...")
            # Try different openpyxl parameters
            params_to_try = [
                {'read_only': True, 'data_only': True},
                {'read_only': False, 'data_only': True},
                {'read_only': True, 'data_only': False}
            ]

            for params in params_to_try:
                try:
                    wb = _oxl_load_workbook(filename=file_path, **params)
                    sheets = wb.sheetnames
                    wb.close()
                    if sheets:
                        print("✅ Success with openpyxl direct read")
                        return sheets
                except Exception:
                    continue

        except Exception as e:
            print(f"  ❌ Openpyxl direct read failed: {str(e)}")

        # Strategy 3: Try win32com if on Windows
        try:
            if os.name == 'nt':  # Windows
                import win32com.client
                print("  Trying win32com...")
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                wb = excel.Workbooks.Open(str(file_path))
                sheets = [sheet.Name for sheet in wb.Sheets]
                wb.Close()
                excel.Quit()
                if sheets:
                    print("✅ Success with win32com")
                    return sheets
        except Exception as e:
            print(f"  ❌ win32com failed: {str(e)}")

        # Strategy 4: Try to read as zip and parse manually for xlsx files
        if file_path.suffix.lower() in ('.xlsx', '.xlsm'):
            try:
                print("  Trying manual zip parsing...")
                import zipfile
                import xml.etree.ElementTree as ET

                with zipfile.ZipFile(file_path, 'r') as z:
                    # Look for workbook.xml to find sheet names
                    with z.open('xl/workbook.xml') as f:
                        tree = ET.parse(f)
                        root = tree.getroot()

                        # Define namespaces
                        ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

                        sheets = []
                        for sheet_elem in root.findall('.//ns:sheet', ns):
                            name = sheet_elem.get('name')
                            if name:
                                sheets.append(name)
                        if sheets:
                            print("✅ Success with manual zip parsing")
                            return sheets
            except Exception as e:
                print(f"  ❌ Manual zip parsing failed: {str(e)}")

        raise RuntimeError(
            "Could not read sheet names with any method.\n"
            "Possible causes:\n"
            "1. File is password protected\n"
            "2. File is corrupted\n"
            "3. File is in an unsupported format\n"
            "4. File is open in another program\n\n"
            "Try:\n"
            "1. Opening and re-saving the file in Excel\n"
            "2. Ensuring the file is not password protected\n"
            "3. Installing: pip install openpyxl xlrd"
        )

    def _read_sheet_df_robust(self, sheet_name):
        """Robust method to read a sheet with multiple fallbacks"""
        file_path = Path(self.excel_file_path)

        # Strategy 1: Try pandas with multiple engines (without calamine)
        engines_to_try = [
            None,  # Let pandas choose
            'openpyxl',
            'xlrd'
        ]

        # Add pyxlsb only for .xlsb files if available
        if file_path.suffix.lower() == '.xlsb':
            try:
                import pyxlsb
                engines_to_try.append('pyxlsb')
            except ImportError:
                pass

        for engine in engines_to_try:
            try:
                print(f"  Reading sheet '{sheet_name}' with engine: {engine}")
                df = pd.read_excel(self.excel_file_path, sheet_name=sheet_name, engine=engine)
                if not df.empty or len(df.columns) > 0:
                    print(f"✅ Success reading with engine: {engine}")
                    return df
            except Exception as e:
                print(f"  ❌ Engine {engine} failed: {str(e)}")
                continue

        # Strategy 2: Manual reading with openpyxl as last resort
        try:
            print(f"  Manual reading sheet '{sheet_name}' with openpyxl...")

            # Try different openpyxl parameters
            params_to_try = [
                {'read_only': True, 'data_only': True},
                {'read_only': False, 'data_only': True},
                {'read_only': True, 'data_only': False}
            ]

            for params in params_to_try:
                try:
                    wb = _oxl_load_workbook(filename=self.excel_file_path, **params)
                    if sheet_name not in wb.sheetnames:
                        continue

                    ws = wb[sheet_name]
                    data = []
                    for row in ws.iter_rows(values_only=True):
                        data.append(row)

                    wb.close()

                    if not data:
                        return pd.DataFrame()

                    # Create DataFrame from data
                    headers = [str(h) if h is not None else f"Column_{i + 1}" for i, h in enumerate(data[0])]
                    rows = data[1:] if len(data) > 1 else []
                    df = pd.DataFrame(rows, columns=headers)
                    print("✅ Success with manual openpyxl reading")
                    return df
                except Exception:
                    continue

        except Exception as e:
            print(f"❌ All methods failed to read sheet '{sheet_name}': {str(e)}")
            raise

    def _ensure_local_copy(self, path):
        """If the workbook is on cloud storage or has issues, create a local copy"""
        try:
            from shutil import copy2
            import tempfile

            p = Path(path).expanduser().resolve()

            if not p.exists():
                return str(p)

            # Check file size
            size = p.stat().st_size
            if size == 0:
                raise ValueError("File is empty (0 bytes)")

            # Check if file is on cloud storage
            cloud_indicators = ['onedrive', 'cloudstorage', 'icloud', 'dropbox', 'google drive']
            file_path_lower = str(p).lower()
            is_cloud = any(indicator in file_path_lower for indicator in cloud_indicators)

            if is_cloud or size < 1024:  # Also copy very small files that might be problematic
                tmp_dir = Path(tempfile.gettempdir())
                tmp_path = tmp_dir / f"excel_processor_{p.name}"
                copy2(str(p), str(tmp_path))
                print(f"📥 Created local copy: {tmp_path}")
                return str(tmp_path)

        except Exception as e:
            print(f"⚠️  Could not create local copy: {e}")

        return str(path)

    def _choose_sheet(self, sheet_names, sheet_type, optional=False):
        """Let user choose which sheet to use"""
        print(f"\nPlease choose which sheet contains your {sheet_type}:")
        for i, name in enumerate(sheet_names, 1):
            print(f"  {i}. {name}")

        if optional:
            print(f"  {len(sheet_names) + 1}. Skip (no {sheet_type} sheet)")

        while True:
            try:
                choice = input(
                    f"Enter your choice (1-{len(sheet_names) + 1 if optional else len(sheet_names)}): ").strip()
                if not choice:
                    continue

                choice_num = int(choice)

                if 1 <= choice_num <= len(sheet_names):
                    return sheet_names[choice_num - 1]
                elif optional and choice_num == len(sheet_names) + 1:
                    return None
                else:
                    print(
                        f"Please enter a number between 1 and {len(sheet_names) + 1 if optional else len(sheet_names)}")
            except ValueError:
                print("Please enter a valid number")

    def _process_customer_data(self, data_sheet):
        """Process and organize data by customer"""
        # Clean column names
        data_sheet.columns = [str(col).strip().title() for col in data_sheet.columns]
        print(f"📊 Processing columns: {list(data_sheet.columns)}")

        # Let user choose which column contains customer names
        customer_col = self._choose_column(data_sheet.columns, "customer names", pref_key="main_customer_column")
        if not customer_col:
            return

        # Group data by customer
        customer_count = 0
        for _, row in data_sheet.iterrows():
            customer_name = str(row.get(customer_col, '')).strip()

            if not customer_name or pd.isna(customer_name) or customer_name == 'nan':
                continue

            # Convert row to dictionary
            row_data = {}
            for col in data_sheet.columns:
                value = row[col]
                if pd.isna(value):
                    value = ""
                row_data[col] = value

            self.customer_data[customer_name].append(row_data)
            customer_count += 1

        print(f"✅ Processed {customer_count} rows for {len(self.customer_data)} customers")

        # Show sample of customers found
        if self.customer_data:
            print("👥 Sample of customers found:")
            for i, customer in enumerate(list(self.customer_data.keys())[:5]):
                print(f"  {i + 1}. {customer} ({len(self.customer_data[customer])} invoices)")

    def _choose_column(self, columns, column_type, pref_key=None):
        """Let user choose which column to use, with optional preference memory."""
        columns = list(columns)

        # Auto-apply saved preference if available and present
        if pref_key:
            saved = self._get_pref(pref_key)
            if saved and saved in columns:
                print(f"\nUsing saved choice for {column_type}: {saved}")
                return saved

        print(f"\nPlease choose which column contains {column_type}:")
        for i, col in enumerate(columns, 1):
            print(f"  {i}. {col}")

        while True:
            try:
                choice = input(f"Enter your choice (1-{len(columns)}): ").strip()
                if not choice:
                    continue

                choice_num = int(choice)

                if 1 <= choice_num <= len(columns):
                    chosen = columns[choice_num - 1]
                    if pref_key:
                        self._save_pref(pref_key, chosen)
                    return chosen
                else:
                    print(f"Please enter a number between 1 and {len(columns)}")
            except ValueError:
                print("Please enter a valid number")

    def _create_emails(self, emails_sheet):
        """Create emails for each customer"""
        try:
            # Get email addresses
            email_addresses = self._get_email_addresses(emails_sheet)

            print(f"📧 Found {len(email_addresses)} customer email addresses")

            print("\nChoose an option:")
            print("1. Send emails automatically (requires SMTP setup)")
            print("2. Save emails as files for manual sending")
            print("3. Display emails in terminal for copy/paste")
            print("4. Export to Excel (xlsxwriter)")

            # New option 5: Save as Outlook drafts (Windows via Outlook; macOS via .eml files)
            print("5. Save emails as Outlook drafts (Windows/macOS Outlook)")

            choice = input("Enter your choice (1-5): ").strip()

            if choice == "1":
                self._send_emails_automatically(email_addresses)
            elif choice == "2":
                self._save_emails_to_files(email_addresses)
            elif choice == "3":
                self._display_emails_in_terminal(email_addresses)
            elif choice == "4":
                self._export_to_xlsxwriter(email_addresses)
            elif choice == "5":
                self._save_emails_to_outlook_drafts(email_addresses)
            else:
                print("Invalid choice. Saving emails to files instead.")
                self._save_emails_to_files(email_addresses)

        except Exception as e:
            print(f"❌ Error creating emails: {str(e)}")
            import traceback
            traceback.print_exc()

    def _get_email_addresses(self, emails_sheet):
        """Extract email addresses from emails sheet"""
        email_addresses = {}

        if emails_sheet.empty:
            print("ℹ️  No emails sheet provided - will prompt for emails during processing")
            return email_addresses

        # Clean column names
        emails_sheet.columns = [str(col).strip().title() for col in emails_sheet.columns]
        print(f"📧 Email sheet columns: {list(emails_sheet.columns)}")

        # Let user choose customer name column
        customer_col = self._choose_column(emails_sheet.columns, "customer names", pref_key="emails_customer_column")
        if not customer_col:
            return email_addresses

        # Let user choose email column
        email_col = self._choose_column(emails_sheet.columns, "email addresses", pref_key="emails_email_column")
        if not email_col:
            return email_addresses

        for _, row in emails_sheet.iterrows():
            customer_name = str(row.get(customer_col, '')).strip()
            email_address = str(row.get(email_col, '')).strip()

            if customer_name and email_address and "@" in email_address:
                # Normalize and store as comma-separated for RFC compliance
                addrs = self._parse_addresses(email_address)
                if addrs:
                    canonical = self._format_addrs_for_mime(addrs)
                    email_addresses[customer_name] = canonical
                    print(f"  ✅ Found email for {customer_name}: {canonical}")
            elif customer_name and email_address:
                print(f"  ⚠️  Invalid email for {customer_name}: {email_address}")

        return email_addresses

    # -------------------- Address helpers --------------------
    def _parse_addresses(self, addr_str):
        """Parse a string of email addresses separated by ';' or ',' into a list."""
        if not addr_str:
            return []
        # Replace semicolons with commas, then split
        tmp = addr_str.replace(';', ',')
        parts = [p.strip() for p in tmp.split(',') if p.strip()]
        # Basic filtering to include only parts containing '@'
        return [p for p in parts if '@' in p]

    def _format_addrs_for_mime(self, addrs):
        return ", ".join(addrs)

    def _format_addrs_for_outlook(self, addrs):
        return "; ".join(addrs)

    def _build_invoice_string(self, invoices):
        """Build formatted string of invoice numbers"""
        invoice_list = []
        for invoice in invoices:
            # Look for invoice number in various possible columns
            inv_num = None
            for col in invoice.keys():
                if any(keyword in col.lower() for keyword in ['invoice', 'inv', '#']):
                    inv_num = invoice.get(col)
                    break

            # If not found, try to find any column that might contain invoice numbers
            if not inv_num:
                for col, value in invoice.items():
                    if value and any(char.isdigit() for char in str(value)):
                        inv_num = value
                        break

            if inv_num and not pd.isna(inv_num) and str(inv_num).strip():
                invoice_list.append(str(inv_num).strip())

        # Build string
        if not invoice_list:
            return "your invoices"
        elif len(invoice_list) == 1:
            return invoice_list[0]
        else:
            return ", ".join(invoice_list[:-1]) + " and " + invoice_list[-1]

    def _create_customer_table(self, invoices):
        """Create HTML table with customer invoice data"""
        if not invoices:
            return "<p>No invoice data available.</p>"

        # Use the actual column names from the data
        columns = list(invoices[0].keys())

        table_html = """
        <table border='1' style='border-collapse: collapse; width: 100%; font-family: Arial, sans-serif; font-size: 12px;'>
            <tr style='background-color: #f2f2f2;'>
        """

        # Add headers
        for col in columns:
            table_html += f"<th style='padding: 8px; text-align: left;'>{col}</th>"

        table_html += "</tr>"

        # Add data rows
        for invoice in invoices:
            table_html += "<tr>"

            for col in columns:
                value = invoice.get(col, "")

                if pd.isna(value):
                    value = ""
                else:
                    value = str(value)

                # Try to detect and format numeric columns
                if any(num_key in col.lower() for num_key in ['amount', 'balance', 'total', 'price', 'due']):
                    try:
                        if value.strip():
                            clean_value = value.replace('$', '').replace(',', '').strip()
                            if clean_value:
                                numeric_value = float(clean_value)
                                value = f"${numeric_value:,.2f}"
                        table_html += f"<td style='padding: 6px; text-align: right;'>{value}</td>"
                    except (ValueError, TypeError):
                        table_html += f"<td style='padding: 6px; text-align: right;'>{value}</td>"
                else:
                    table_html += f"<td style='padding: 6px;'>{value}</td>"

            table_html += "</tr>"

        table_html += "</table>"
        return table_html

    def _calculate_total_amount(self, invoices):
        """Calculate total amount from invoices"""
        total = 0
        for invoice in invoices:
            # Look for amount/balance columns
            for col in invoice.keys():
                if any(keyword in col.lower() for keyword in ['balance', 'amount', 'total', 'due']):
                    try:
                        value = invoice.get(col, 0)
                        if pd.isna(value):
                            value = 0
                        total += float(str(value).replace('$', '').replace(',', ''))
                        break
                    except (ValueError, TypeError):
                        pass
        return total

    def _export_to_xlsxwriter(self, email_addresses):
        """Export per-customer data and email drafts into an .xlsx using xlsxwriter."""
        try:
            try:
                import xlsxwriter
            except ImportError:
                print("❌ xlsxwriter is not installed. Please run: pip install XlsxWriter")
                return

            src_path = Path(self.excel_file_path)
            out_dir = src_path.parent if src_path.parent else Path.cwd()
            out_name = src_path.stem + "_emails_export.xlsx"
            out_path = out_dir / out_name

            # If file exists, attempt to remove to avoid permission issues
            try:
                if out_path.exists():
                    out_path.unlink()
            except Exception:
                pass

            workbook = xlsxwriter.Workbook(str(out_path))

            # Formats
            fmt_bold = workbook.add_format({'bold': True})
            fmt_hdr = workbook.add_format({'bold': True, 'bg_color': '#F2F2F2', 'border': 1})
            fmt_text = workbook.add_format({'border': 1})
            fmt_curr = workbook.add_format({'num_format': '$#,##0.00', 'border': 1, 'align': 'right'})
            fmt_right = workbook.add_format({'align': 'right', 'border': 1})

            # Helper to sanitize sheet names
            def _safe_sheet_name(name, existing):
                s = str(name).strip() or "Sheet"
                # Replace invalid characters
                for ch in [':', '\\', '/', '?', '*', '[', ']']:
                    s = s.replace(ch, ' ')
                s = s[:31]
                base = s
                i = 1
                while s in existing:
                    suffix = f"_{i}"
                    s = (base[:31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else base + suffix
                    i += 1
                existing.add(s)
                return s

            existing_names = set()

            # Summary sheet
            ws_summary = workbook.add_worksheet(_safe_sheet_name('Summary', existing_names))
            summary_headers = ["Customer", "Email", "Subject", "Total", "Invoice List", "Invoice Count"]
            for col, h in enumerate(summary_headers):
                ws_summary.write(0, col, h, fmt_hdr)

            # Emails sheet
            ws_emails = workbook.add_worksheet(_safe_sheet_name('Emails', existing_names))
            emails_headers = ["Customer", "To", "Cc", "Subject", "HTML Body"]
            for col, h in enumerate(emails_headers):
                ws_emails.write(0, col, h, fmt_hdr)

            # CC addresses
            default_cc = "executive.admin@ranaanalytics.com; michael.wells@ranaanalytics.com"

            summary_row = 1
            emails_row = 1

            # Build per-customer sheets
            for customer, invoices in self.customer_data.items():
                if not invoices:
                    continue

                # Determine columns order from first invoice
                columns = list(invoices[0].keys())

                # Write customer sheet
                ws_name = _safe_sheet_name(customer, existing_names)
                ws_cust = workbook.add_worksheet(ws_name)

                # Headers
                for col_idx, col_name in enumerate(columns):
                    ws_cust.write(0, col_idx, col_name, fmt_hdr)

                # Rows
                for r_idx, inv in enumerate(invoices, start=1):
                    for c_idx, col_name in enumerate(columns):
                        val = inv.get(col_name, "")
                        if val is None or (isinstance(val, float) and pd.isna(val)):
                            val = ""
                        # Check if column looks like currency
                        if any(keyword in str(col_name).lower() for keyword in
                               ['amount', 'balance', 'total', 'price', 'due']):
                            try:
                                if isinstance(val, str):
                                    clean = val.replace('$', '').replace(',', '').strip()
                                    val_num = float(clean) if clean else 0.0
                                else:
                                    val_num = float(val)
                                ws_cust.write_number(r_idx, c_idx, val_num, fmt_curr)
                            except Exception:
                                ws_cust.write(r_idx, c_idx, str(val), fmt_right)
                        else:
                            ws_cust.write(r_idx, c_idx, val, fmt_text)

                # Autosize columns
                for c_idx, col_name in enumerate(columns):
                    max_len = max([len(str(col_name))] + [len(str(inv.get(col_name, ""))) for inv in invoices])
                    ws_cust.set_column(c_idx, c_idx, min(max(10, max_len + 2), 60))

                # Compute totals and email pieces
                total_amount = self._calculate_total_amount(invoices)

                # Try to find invoice number column
                inv_col_candidates = [c for c in columns if 'invoice' in str(c).lower() or 'inv' in str(c).lower()]
                inv_col = inv_col_candidates[0] if inv_col_candidates else None
                invoice_values = []
                if inv_col:
                    for inv in invoices:
                        v = inv.get(inv_col, "")
                        v = "" if (isinstance(v, float) and pd.isna(v)) else v
                        if str(v).strip():
                            invoice_values.append(str(v).strip())

                invoice_list_text = (
                    invoice_values[0] if len(invoice_values) == 1 else
                    (", ".join(invoice_values[:-1]) + " and " + invoice_values[-1]) if len(invoice_values) > 1 else
                    "your invoices"
                )

                subject = f"Rana Analytics - {customer} Account Overdue Notice"

                # Create HTML table and body
                table_html = self._create_customer_table(invoices)
                html_body = (
                        "<html><body>"
                        f"Dear {customer},<br><br>"
                        f"This is a friendly reminder that your invoice numbers {invoice_list_text} are now overdue. "
                        f"As per our records, the amount owing is <strong>${total_amount:,.2f}</strong>.<br><br>"
                        f"<h3>Invoice Details:</h3>" + table_html + "<br>"
                                                                    "Please kindly send payment through the payment modes we have listed below.<br><br>"
                                                                    "We appreciate your quick attention to this matter and look forward to our continued partnership.<br><br>"
                                                                    "Regards,<br><br>"
                                                                    "<strong>Rana Analytics, LLC</strong><br>"
                                                                    "Accounts Receivable<br><br>"
                                                                    "<h3>Payment Information:</h3>"
                                                                    "Bank Name: Texas First Bank<br>"
                                                                    "Account Name: Rana Analytics LLC<br>"
                                                                    "ACH Routing #: 113110256<br>"
                                                                    "Account #: 10420917<br>"
                                                                    "Type of Account: Checking<br>"
                                                                    "</body></html>"
                )

                # Append to Summary
                ws_summary.write(summary_row, 0, customer, fmt_text)
                ws_summary.write(summary_row, 1, email_addresses.get(customer, ""), fmt_text)
                ws_summary.write(summary_row, 2, subject, fmt_text)
                ws_summary.write_number(summary_row, 3, total_amount, fmt_curr)
                ws_summary.write(summary_row, 4, ", ".join(invoice_values), fmt_text)
                ws_summary.write_number(summary_row, 5, len(invoices), fmt_right)
                summary_row += 1

                # Append to Emails sheet
                ws_emails.write(emails_row, 0, customer, fmt_text)
                ws_emails.write(emails_row, 1, email_addresses.get(customer, ""), fmt_text)
                ws_emails.write(emails_row, 2, default_cc, fmt_text)
                ws_emails.write(emails_row, 3, subject, fmt_text)
                ws_emails.write_string(emails_row, 4, html_body)
                emails_row += 1

            # Autosize summary columns
            for c in range(6):
                ws_summary.set_column(c, c, 20 if c in (0, 1, 2) else 18)

            workbook.close()
            print(f"✅ Exported email workbook via xlsxwriter: {out_path}")
        except Exception as e:
            print(f"❌ Failed to export using xlsxwriter: {e}")
            import traceback
            traceback.print_exc()

    def _create_email_body(self, customer_name, invoice_numbers, total_amount, customer_table):
        """Create the complete email HTML body"""
        total_formatted = f"${total_amount:,.2f}"

        return f"""
<html>
<body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
<div style="max-width: 800px; margin: 0 auto;">
    <p>Dear {customer_name},</p>

    <p>This is a friendly reminder that your invoice numbers {invoice_numbers} are now overdue. 
    As per our records, the amount owing is <strong>{total_formatted}</strong>.</p>

    <h3 style="color: #2c3e50;">Invoice Details:</h3>
    {customer_table}

    <p>Please kindly send payment through the payment modes we have listed below.</p>

    <p>We appreciate your quick attention to this matter and look forward to our continued partnership.</p>

    <p>Regards,</p>

    <p><strong>Rana Analytics, LLC</strong><br>
    Accounts Receivable</p>

    <h3 style="color: #2c3e50;">Payment Information:</h3>
    <ul>
        <li>Bank Name: Texas First Bank</li>
        <li>Account Name: Rana Analytics LLC</li>
        <li>ACH Routing #: 113110256</li>
        <li>Account #: 10420917</li>
        <li>Type of Account: Checking</li>
    </ul>
</div>
</body>
</html>
        """

    def _create_plain_text_body(self, customer_name, invoice_numbers, total_amount, invoices):
        """Create plain text version"""
        total_formatted = f"${total_amount:,.2f}"

        text = f"""Dear {customer_name},

This is a friendly reminder that your invoice numbers {invoice_numbers} are now overdue. 
As per our records, the amount owing is {total_formatted}.

INVOICE DETAILS:
"""
        if invoices:
            headers = list(invoices[0].keys())
            text += " | ".join(headers) + "\n"
            text += "-" * (len(headers) * 15) + "\n"

            for invoice in invoices:
                row = []
                for header in headers:
                    value = invoice.get(header, "")
                    if pd.isna(value):
                        value = ""
                    row.append(str(value))
                text += " | ".join(row) + "\n"

        text += f"""
Please kindly send payment through the payment modes we have listed below.

We appreciate your quick attention to this matter and look forward to our continued partnership.

Regards,

Rana Analytics, LLC
Accounts Receivable

PAYMENT INFORMATION:
Bank Name: Texas First Bank
Account Name: Rana Analytics LLC
ACH Routing #: 113110256
Account #: 10420917
Type of Account: Checking
"""
        return text

    def _send_emails_automatically(self, email_addresses):
        """Send emails automatically using SMTP"""
        print("\n=== SMTP Email Setup ===")
        smtp_server = input("SMTP Server (e.g., smtp.gmail.com): ").strip()
        smtp_port = int(input("SMTP Port (e.g., 587): ").strip())
        email_from = input("Your Email: ").strip()
        email_password = input("Your Email Password/App Password: ").strip()

        try:
            for customer_name, invoices in self.customer_data.items():
                customer_email = email_addresses.get(customer_name)

                if not customer_email:
                    customer_email = input(f"Enter email address for {customer_name}: ").strip()
                    if not customer_email:
                        print(f"⚠️  Skipping {customer_name}")
                        continue

                # Prepare email content
                invoice_numbers = self._build_invoice_string(invoices)
                total_amount = self._calculate_total_amount(invoices)

                html_body = self._create_email_body(customer_name, invoice_numbers, total_amount,
                                                    self._create_customer_table(invoices))
                plain_text_body = self._create_plain_text_body(customer_name, invoice_numbers,
                                                               total_amount, invoices)

                # Create message
                DEFAULT_CC = ["executive.admin@ranaanalytics.com", "michael.wells@ranaanalytics.com"]
                msg = MIMEMultipart('alternative')
                msg['Subject'] = f"Rana Analytics - {customer_name} Account Overdue Notice"
                msg['From'] = email_from
                # Normalize recipients for SMTP/MIME (comma-separated)
                to_addrs = self._parse_addresses(customer_email)
                msg['To'] = self._format_addrs_for_mime(to_addrs) if to_addrs else customer_email
                msg['Cc'] = self._format_addrs_for_mime(DEFAULT_CC)

                msg.attach(MIMEText(plain_text_body, 'plain'))
                msg.attach(MIMEText(html_body, 'html'))

                # Send email
                with smtplib.SMTP(smtp_server, smtp_port) as server:
                    server.starttls()
                    server.login(email_from, email_password)
                    server.send_message(msg)

                print(f"✅ Email sent to: {customer_name} ({customer_email})")

        except Exception as e:
            print(f"❌ Error sending emails: {str(e)}")

    def _save_emails_to_outlook_drafts(self, email_addresses):
        """Create Outlook draft emails.

        - On Windows: uses Outlook COM automation to save directly into Drafts.
        - On macOS: writes .eml files compatible with Microsoft Outlook and can auto-open them.
        """
        print("\n💼 Creating Outlook draft emails...")

        import sys
        platform = sys.platform

        # Windows path: COM automation into Drafts
        if platform.startswith('win'):
            try:
                import win32com.client  # type: ignore
                from win32com.client import constants  # type: ignore
            except ImportError:
                print("❌ python -m pip install pywin32 is required for Outlook integration on Windows.")
                return
            except Exception as e:
                print(f"❌ Unable to initialize Outlook integration: {e}")
                return

            try:
                outlook = win32com.client.Dispatch("Outlook.Application")
                namespace = outlook.GetNamespace("MAPI")
                drafts_folder = namespace.GetDefaultFolder(constants.olFolderDrafts)
            except Exception as e:
                print(f"❌ Could not access Outlook. Ensure Outlook is installed and configured. Details: {e}")
                return

            created = 0
            skipped = 0
            DEFAULT_CC = ["executive.admin@ranaanalytics.com", "michael.wells@ranaanalytics.com"]
            default_from = self._get_pref("default_from_email")
            for customer_name, invoices in self.customer_data.items():
                try:
                    customer_email = email_addresses.get(customer_name)
                    if not customer_email:
                        customer_email = input(f"Enter email address for {customer_name}: ").strip()
                        if not customer_email:
                            print(f"⚠️  Skipping {customer_name} (no email)")
                            skipped += 1
                            continue

                    invoice_numbers = self._build_invoice_string(invoices)
                    total_amount = self._calculate_total_amount(invoices)
                    html_body = self._create_email_body(customer_name, invoice_numbers, total_amount,
                                                        self._create_customer_table(invoices))
                    plain_text_body = self._create_plain_text_body(customer_name, invoice_numbers,
                                                                   total_amount, invoices)

                    mail = outlook.CreateItem(0)  # 0 = olMailItem
                    subject = f"Rana Analytics - {customer_name} Account Overdue Notice"
                    mail.Subject = subject
                    # Normalize recipients for Outlook (semicolon-separated)
                    to_addrs = self._parse_addresses(customer_email)
                    mail.To = self._format_addrs_for_outlook(to_addrs) if to_addrs else customer_email
                    mail.CC = self._format_addrs_for_outlook(DEFAULT_CC)
                    # Attempt to set From (requires proper permissions in Outlook profile)
                    if default_from:
                        try:
                            mail.SentOnBehalfOfName = default_from
                        except Exception:
                            pass
                    mail.Body = plain_text_body
                    mail.HTMLBody = html_body
                    mail.Save()  # Saves to Drafts by default

                    created += 1
                    print(f"✅ Draft created for: {customer_name} ({customer_email})")
                except Exception as e:
                    print(f"❌ Failed to create draft for {customer_name}: {e}")
                    skipped += 1

            try:
                print(f"\n📨 Draft creation summary: {created} created, {skipped} skipped.")
                drafts_count = len(drafts_folder.Items)
                print(f"📁 Outlook Drafts folder now contains approximately {drafts_count} items.")
            except Exception:
                pass
            return

        # macOS path: write .eml files and optionally open in Outlook
        if platform == 'darwin':
            from pathlib import Path
            from email.mime.multipart import MIMEMultipart
            from email.mime.text import MIMEText
            from email import policy
            import subprocess

            out_dir = Path("Customer_Emails") / "Outlook_Drafts_EML"
            out_dir.mkdir(parents=True, exist_ok=True)

            created = 0
            skipped = 0

            # Optionally allow user to set a default From for EML drafts (remembered)
            default_from = self._get_pref("default_from_email")
            if not default_from:
                try:
                    resp = input("Enter a default From email to use in drafts (press Enter to skip): ").strip()
                    if resp:
                        default_from = resp
                        self._save_pref("default_from_email", default_from)
                except Exception:
                    pass

            for customer_name, invoices in self.customer_data.items():
                try:
                    customer_email = email_addresses.get(customer_name)
                    if not customer_email:
                        customer_email = input(f"Enter email address for {customer_name}: ").strip()
                        if not customer_email:
                            print(f"⚠️  Skipping {customer_name} (no email)")
                            skipped += 1
                            continue

                    invoice_numbers = self._build_invoice_string(invoices)
                    total_amount = self._calculate_total_amount(invoices)
                    html_body = self._create_email_body(customer_name, invoice_numbers, total_amount,
                                                        self._create_customer_table(invoices))
                    plain_text_body = self._create_plain_text_body(customer_name, invoice_numbers,
                                                                   total_amount, invoices)

                    DEFAULT_CC = ["executive.admin@ranaanalytics.com", "michael.wells@ranaanalytics.com"]
                    msg = MIMEMultipart('alternative')
                    subject = f"Rana Analytics - {customer_name} Account Overdue Notice"
                    msg['Subject'] = subject
                    to_addrs = self._parse_addresses(customer_email)
                    msg['To'] = self._format_addrs_for_mime(to_addrs) if to_addrs else customer_email
                    msg['Cc'] = self._format_addrs_for_mime(DEFAULT_CC)
                    if default_from:
                        msg['From'] = default_from

                    msg.attach(MIMEText(plain_text_body, 'plain'))
                    msg.attach(MIMEText(html_body, 'html'))

                    # Clean filename
                    clean_name = "".join(c for c in customer_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
                    clean_name = clean_name.replace(' ', '_')
                    file_path = out_dir / f"{clean_name}.eml"

                    with open(file_path, 'wb') as f:
                        f.write(msg.as_bytes(policy=policy.SMTP))

                    created += 1
                    print(f"✅ EML draft saved for: {customer_name} → {file_path}")
                except Exception as e:
                    print(f"❌ Failed to create EML for {customer_name}: {e}")
                    skipped += 1

            print(f"\n📨 EML creation summary: {created} created, {skipped} skipped.")
            if created:
                try:
                    open_now = input("Open the .eml drafts in Microsoft Outlook now? (y/N): ").strip().lower()
                    if open_now == 'y':
                        # Attempt to open all .eml files in Outlook. This works with both Classic and New Outlook.
                        for eml_file in sorted(out_dir.glob('*.eml')):
                            try:
                                subprocess.run(["open", "-a", "Microsoft Outlook", str(eml_file)], check=False)
                            except Exception:
                                # Best-effort; continue with others
                                pass
                        print("✅ Requested Outlook to open all .eml drafts. In Outlook, click Save to place them into your Drafts folder.")
                except Exception:
                    pass
            return

        # Other platforms
        print("❌ Drafting via Outlook is supported on Windows (Outlook Desktop) and macOS (via .eml files). Your platform is not supported.")

    def _save_emails_to_files(self, email_addresses):
        """Save emails as HTML files for manual sending"""
        output_dir = Path("Customer_Emails")
        output_dir.mkdir(exist_ok=True)

        print(f"\n💾 Saving emails to: {output_dir.absolute()}")

        for customer_name, invoices in self.customer_data.items():
            customer_email = email_addresses.get(customer_name)
            if not customer_email:
                customer_email = input(f"Enter email address for {customer_name}: ").strip()
                if not customer_email:
                    print(f"⚠️  Skipping {customer_name}")
                    continue

            # Prepare email content
            invoice_numbers = self._build_invoice_string(invoices)
            total_amount = self._calculate_total_amount(invoices)

            html_body = self._create_email_body(customer_name, invoice_numbers, total_amount,
                                                self._create_customer_table(invoices))

            # Clean filename
            clean_name = "".join(c for c in customer_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
            clean_name = clean_name.replace(' ', '_')
            filename = output_dir / f"{clean_name}_email.html"

            with open(filename, 'w', encoding='utf-8') as f:
                f.write(html_body)

            print(f"✅ Email saved for: {customer_name}")
            print(f"   📄 File: {filename.name}")
            print(f"   📧 To: {customer_email}")
            print(f"   💰 Total Amount: ${total_amount:,.2f}")
            print(f"   📋 Invoices: {invoice_numbers}\n")

        print(f"\n🎉 All emails saved in: {output_dir.absolute()}")
        print("You can open these HTML files in your browser and copy/paste the content into your email client.")

    def _display_emails_in_terminal(self, email_addresses):
        """Display email content in terminal for copy/paste"""
        for customer_name, invoices in self.customer_data.items():
            customer_email = email_addresses.get(customer_name)
            if not customer_email:
                customer_email = input(f"Enter email address for {customer_name}: ").strip()
                if not customer_email:
                    print(f"⚠️  Skipping {customer_name}")
                    continue

            # Prepare email content
            invoice_numbers = self._build_invoice_string(invoices)
            total_amount = self._calculate_total_amount(invoices)

            plain_text_body = self._create_plain_text_body(customer_name, invoice_numbers,
                                                           total_amount, invoices)

            print(f"\n{'=' * 80}")
            print(f"EMAIL FOR: {customer_name}")
            print(f"TO: {customer_email}")
            print(f"CC: executive.admin@ranaanalytics.com, michael.wells@ranaanalytics.com")
            print(f"SUBJECT: Rana Analytics - {customer_name} Account Overdue Notice")
            print(f"{'=' * 80}")
            print(plain_text_body)
            print(f"{'=' * 80}\n")

            input("Press Enter to continue to next customer...")


def main():
    """Main function to process customer emails from Excel file"""
    print("=== Customer Email Processor ===")
    print("This tool processes customer data and creates overdue payment emails.")
    print()

    # Install required packages if not already installed
    required_packages = ['pandas', 'openpyxl']
    missing_packages = []

    for package in required_packages:
        try:
            __import__(package)
        except ImportError:
            missing_packages.append(package)

    if missing_packages:
        print(f"Installing required packages: {missing_packages}...")
        import subprocess
        import sys
        subprocess.check_call([sys.executable, "-m", "pip", "install"] + missing_packages)
        print("✅ Packages installed successfully!")

    # Get Excel file path
    # Provide a convenient default known location mentioned by you.
    default_excel_path = "/Users/pauldesoloc/Downloads/Ranas.xlsx"
    prompt = "Enter the path to your Excel file"
    if os.path.exists(default_excel_path):
        prompt += f" (Press Enter to use default: {default_excel_path})"
    prompt += ": "

    user_input_path = input(prompt).strip()

    if user_input_path:
        excel_file = user_input_path
        if not os.path.exists(excel_file):
            # If the provided path doesn't exist but the default does, use the default transparently
            if os.path.exists(default_excel_path):
                print(f"⚠️  Provided path not found. Using default: {default_excel_path}")
                excel_file = default_excel_path
            else:
                print("❌ File not found! Please check the path and try again.")
                return
    else:
        # Empty input: try default if present
        if os.path.exists(default_excel_path):
            excel_file = default_excel_path
            print(f"✅ Using default Excel file: {excel_file}")
        else:
            print("❌ No path provided and default file not found. Please run again and provide a valid path.")
            return

    # Process the file
    processor = CustomerEmailProcessor(excel_file)
    processor.separate_data_per_customer()


if __name__ == "__main__":
    main()